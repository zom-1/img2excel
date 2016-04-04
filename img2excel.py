#!/usr/bin/env python
# coding: utf-8\
import os
import argparse
from PIL import Image  # Pillow or PIL
from openpyxl import Workbook  # library to r/w excel -> https://openpyxl.readthedocs.org/
from openpyxl.styles import PatternFill
__author__ = "zom-1"
__version__ = "0.5.1"
__date__ = "2016/04/04"


def img2Excel(imgFile, excelFile, columns):
    ''' make a Excel mosaic picture from image '''
    try:
        img = Image.open(imgFile)
    except IOError:
        print 'cannot open', imgFile
        quit()

    # make small image
    orgWidth, orgHight = img.size
    smlWidth = columns
    smlHeight = smlWidth*orgHight/orgWidth
    smlImg = img.resize((smlWidth, smlHeight), Image.ANTIALIAS)

    # make excel file
    wb = Workbook()  # new excel, TODO: error check
    ws1 = wb.active
    ws1.title = excelFile
    for y in range(0, smlHeight):
        for x in range(0, smlWidth):
            r, g, b = smlImg.getpixel((x, y))
            c = "{0:0>2X}{1:0>2X}{2:0>2X}".format(r, g, b)  # hex, ex)FF88CC
            ws1.cell(row=y+1, column=x+1, value=" ").fill = \
                PatternFill(fill_type='solid', start_color=c, end_color=c)

    # n2A1 : column num to A1 style : 1->'A',2->'B'... 27->'AA',28->'AB'
    def n2A1(n): return (chr(((n-1)//26)+64) if n > 26 else '') + chr((n-1) % 26 + 65)
    for n in range(1, smlWidth+1):
        ws1.column_dimensions[n2A1(n)].width = 2.5  # set column width, TODO: better way ???
    wb.save(excelFile)

if __name__ == '__main__':
    # args parse
    parser = argparse.ArgumentParser()
    parser.add_argument("-c", "--columns", default=50, type=int, help="columns of Excel")
    parser.add_argument("-e", "--excel_file",  default="",  help="Excel file")
    parser.add_argument("image_file", help="output a Excel of a given image")
    args = parser.parse_args()
    if (args.excel_file):  # if empty, use image basename+'xlsx'
        excelFile = args.excel_file
    else:
        excelFile = os.path.splitext(os.path.basename(args.image_file))[0]+".xlsx"

    # exec
    img2Excel(args.image_file, excelFile, args.columns)
